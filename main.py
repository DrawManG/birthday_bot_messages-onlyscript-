from datetime import datetime, date, time
import openpyxl

#Массивы для заготовок
mouth_data = []
fio_data = []
data_data = []
phone_data = []

#Количество строк
row_max = 73

#открытие ексельки
book = openpyxl.load_workbook('HappyDay.xlsx')

sheet = book.active

mouth_head = sheet.cell(row=1,column=2) #mouth
fio_head = sheet.cell(row=1,column=3) #fio
data_head = sheet.cell(row=1, column=4) #data rojd
phone_head = sheet.cell(row=1, column=6) #number

#чтение
i = 0
while i < row_max:
    if not str(sheet.cell(row=1+i,column=2).value) == 'None':
        mouth_data.append(sheet.cell(row=1+i,column=2).value)
    if not str(sheet.cell(row=1+i,column=3).value) == 'None':
        fio_data.append(sheet.cell(row=1+i,column=3).value)
    if not str(sheet.cell(row=1+i,column=4).value) == 'None':
        data_data.append(str(sheet.cell(row=1+i, column=4).value))
    if not str(sheet.cell(row=1+i,column=6).value) == 'None':
        phone_data.append(sheet.cell(row=1+i, column=6).value)
    i = i + 1

#Дата сейчас
Now_month = datetime(2021,datetime.today().month,datetime.today().day) # TODO МЕНЯЕТСЯ ЛИ ГОД ИЛИ ТАК ОСТАВИТЬ??
try:
    #Если находит ДР
  happy_day = data_data.index(str(Now_month))

  print('Сегодня день рождения у: ' + fio_data[happy_day] + ', ' + 'его номер телефона для поздравлений:' + phone_data[happy_day])
except:
    #Если не находит ДР
    print('Сегодня нету дня рождения...')